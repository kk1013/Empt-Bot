import discord
import asyncio
import datetime
import random
import openpyxl
import bs4
import urllib
import request
from discord import Member
from discord.ext import commands
from urllib.request import urlopen, Request
import urllib
import urllib.request
import os
import sys
import json
import time
import datetime


client = discord.Client()

@client.event
async def on_ready():
    print("login")
    print(client.user.name)
    print(client.user.id)
    print("--------------------")
    await client.change_presence(game=discord.Game(name='엠트가 시켜서 대화중...', type=1))

@client.event
async def on_message(message):
    if message.content.startswith('!프로필'):
        date = datetime.datetime.utcfromtimestamp(((int(message.author.id) >> 22) + 1420070400000) / 1000)
        embed = discord.Embed(coler=0x00ffbb)
        embed.add_field(name="사용자 이름", value=message.author.name, inline=True)
        embed.add_field(name="서버내 이름", value=message.author.display_name, inline=True)
        embed.add_field(name="계정생성일", value=str(date.year) + "년" + str(date.month) + "월" + str(date.day) + "일", inline=True)
        embed.set_thumbnail(url=message.author.avatar_url)
        await client.send_message(message.channel, embed=embed)

    if message.content.startswith("!안녕"):
        await client.send_message(message.channel, "안녕하세요")

    if message.content.startswith("!롤하실분"):
        await client.send_message(message.channel, "@everyone 롤하실분들!! 구합니다!")

    if message.content.startswith("!바보"):
        await client.send_message(message.channel, "응 니달이\n애니\n미스포츈 좋더라")

    if message.content.startswith("엠트"):
        await client.send_message(message.channel, "지금 엠트님은 수면 중이니 급한거 있으시면 개인으로 보내주세요.")

    if message.content.startswith("!사랑해"):
        await client.send_message(message.channel, "나는 여자가 좋아 ")

    if message.content.startswith("!팀나누기"):
        team = message.content[6:]
        peopleteam = team.split("/")
        people = peopleteam[0]
        team = peopleteam[1]
        person = people.split(" ")
        teamname = team.split(" ")
        random.shuffle(teamname)
        for i in range(0, len(person)):
            await client.send_message(message.channel, person[i] + "---->" + teamname[i])

    if message.content.startswith("!서버"):

        list = []
        for server in client.servers:
            list.append(server.name)
        await client.send_message(message.channel, "\n".join(list))


    if message.content.startswith("!투표"):
        vote = message.content[4:].split("/")
        await client.send_message(message.channel, "★투표 - " + vote[0])
        for i in range(1, len(vote)):
            choose = await client.send_message(message.channel, "```" + vote[i] + "```")
            await client.add_reaction(choose, '☺')

    if message.content.startswith("!명령어"):
        await client.send_message(message.channel, "``` 1. !안녕```")
        await client.send_message(message.channel, "``` 2. !롤하실분```")
        await client.send_message(message.channel, "``` 3. !바보```")
        await client.send_message(message.channel, "``` 4. 엠트```")
        await client.send_message(message.channel, "``` 5. !사랑해```")
        await client.send_message(message.channel, "``` 6. !프로필```")
        await client.send_message(message.channel, "``` 7. !서버```")

    if "바보" in message.content or "멍청이" in message.content:
        file = openpyxl.load.workbook("경고.xlsx")
        sheet = file.active
        member = discord.utils.get(client.get_all_members(), id="553935362716729356")
        for i in range(1, 32):
            if str(sheet["A" + str(i)].value) == str(message.author.id):
                sheet["B" + str(i)].value == int(sheet["B" + str(i)].value) + 1
                if int(sheet["B" + str(i)].value) == 3:
                    await client.ban(member, 1)
                break

            if str(sheet["A" + str(i)].value) == "-":
                sheet["A" + str(i)].value = str(message.author.id)
                sheet["B" + str(i)].value = 1
                break
        file.save("경고.xlsx")
        await client.send_message(message.channel, "경고를 받았습니다. 단어선택에 주의해주세요.")

    if message.content.startswith("!기억"):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        learn = message.content.split(" ")
        for i in range (1, 62):
            if sheet["A" + str(i)].value == "-":
                sheet["A" + str(i)].value = learn[1]
                sheet["B" + str(i)].value = learn[2]
                await client.send_message(message.channel, "단어가 학습되었습니다.")
                break
        file.save("기억.xlsx")

    if message.content.startswith("!학습"):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        learn = message.content.split(" ")
        for i in range(1, 62):
            if sheet["A" + str(i)].value == "-":
                sheet["A" + str(i)].value = learn[1]
                sheet["B" + str(i)].value = learn[2]
                break
        file.save("기억.xlsx")
        await client.send_message(message.channel, "단어가 학습되었습니다.")

    if message.content.startswith("!삭제"):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        memory = message.content.split(" ")
        for i in range(1, 62):
            if sheet["A" + str(i)].value == str(memory[1]):
                sheet["A" + str(i)].value = "-"
                sheet["B" + str(i)].value = "-"
                await client.send_message(message.channel, "기억이 삭제 되었습니다.")
                file.save("기억.xlsx")
                break

    if message.content.startswith("!기억삭제"):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        memory = message.content.split(" ")
        for i in range(1, 62):
            if sheet["A" + str(i)].value == str(memory[1]):
                sheet["A" + str(i)].value = "-"
                sheet["B" + str(i)].value = "-"
                await client.send_message(message.channel, "기억이 삭제 되었습니다.")
                file.save("기억.xlsx")
                break

    if message.content.startswith("!불러오기"):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        memory = message.content.split(" ")
        for i in range(1, 62):
            if sheet["A" + str(i)].value == memory[1]:
                await client.send_message(message.channel, sheet["B" + str(i)].value)
                break

    if message.content.startswith("!배그솔로"):

        learn = message.content.split(" ")
        location = learn[1]
        enc_location = urllib.parse.quote(location)
        url = "https://dak.gg/profile/" + enc_location
        html = urllib.request.urlopen(url)
        bsObj = bs4.BeautifulSoup(html, "html.parser")
        solo1 = bsObj.find("div", {"class": "overview"})
        solo2 = solo1.text
        solo3 = solo2.strip()
        channel = message.channel
        embed = discord.Embed(
            title='배그솔로 정보',
            description='배그솔로 정보입니다.',
            colour=discord.Colour.green())
        if solo3 == "No record":
            print("솔로 경기가 없습니다.")
            embed.add_field(name='배그를 한판이라도 해주세요', value='솔로 경기 전적이 없습니다..', inline=False)
            await client.send_message(channel, embed=embed)

        else:
            solo4 = solo1.find("span", {"class": "value"})
            soloratting = solo4.text  # -------솔로레이팅---------
            solorank0_1 = solo1.find("div", {"class": "grade-info"})
            solorank0_2 = solorank0_1.text
            solorank = solorank0_2.strip()  # -------랭크(그마,브론즈)---------

            print("레이팅 : " + soloratting)
            print("등급 : " + solorank)
            print("")
            embed.add_field(name='레이팅', value=soloratting, inline=False)
            embed.add_field(name='등급', value=solorank, inline=False)

            soloKD1 = bsObj.find("div", {"class": "kd stats-item stats-top-graph"})
            soloKD2 = soloKD1.find("p", {"class": "value"})
            soloKD3 = soloKD2.text
            soloKD = soloKD3.strip()  # -------킬뎃(2.0---------
            soloSky1 = soloKD1.find("span", {"class": "top"})
            soloSky2 = soloSky1.text  # -------상위10.24%---------

            print("킬뎃 : " + soloKD)
            print("킬뎃상위 : " + soloSky2)
            print("")
            embed.add_field(name='킬뎃,킬뎃상위', value=soloKD + " " + soloSky2, inline=False)
            # embed.add_field(name='킬뎃상위', value=soloSky2, inline=False)

            soloWinRat1 = bsObj.find("div", {"class": "stats"})  # 박스
            soloWinRat2 = soloWinRat1.find("div", {"class": "winratio stats-item stats-top-graph"})
            soloWinRat3 = soloWinRat2.find("p", {"class": "value"})
            soloWinRat = soloWinRat3.text.strip()  # -------승률---------
            soloWinRatSky1 = soloWinRat2.find("span", {"class": "top"})
            soloWinRatSky = soloWinRatSky1.text.strip()  # -------상위?%---------

            print("승률 : " + soloWinRat)
            print("승률상위 : " + soloWinRatSky)
            print("")
            embed.add_field(name='승률,승률상위', value=soloWinRat + " " + soloWinRatSky, inline=False)
            # embed.add_field(name='승률상위', value=soloWinRatSky, inline=False)

            soloHead1 = soloWinRat1.find("div", {"class": "headshots stats-item stats-top-graph"})
            soloHead2 = soloHead1.find("p", {"class": "value"})
            soloHead = soloHead2.text.strip()  # -------헤드샷---------
            soloHeadSky1 = soloHead1.find("span", {"class": "top"})
            soloHeadSky = soloHeadSky1.text.strip()  # # -------상위?%---------

            print("헤드샷 : " + soloHead)
            print("헤드샷상위 : " + soloHeadSky)
            print("")
            embed.add_field(name='헤드샷,헤드샷상위', value=soloHead + " " + soloHeadSky, inline=False)
            # embed.add_field(name='헤드샷상위', value=soloHeadSky, inline=False)
            await client.send_message(channel, embed=embed)

    if message.content.startswith("!배그듀오"):

        learn = message.content.split(" ")
        location = learn[1]
        enc_location = urllib.parse.quote(location)
        url = "https://dak.gg/profile/" + enc_location
        html = urllib.request.urlopen(url)
        bsObj = bs4.BeautifulSoup(html, "html.parser")
        duoCenter1 = bsObj.find("section", {"class": "duo modeItem"})
        duoRecord1 = duoCenter1.find("div", {"class": "overview"})
        duoRecord = duoRecord1.text.strip()  # ----기록이없습니다 문구----
        print(duoRecord)
        channel = message.channel
        embed = discord.Embed(
            title='배그듀오 정보',
            description='배그듀오 정보입니다.',
            colour=discord.Colour.green())
        if duoRecord == 'No record':
            print('듀오 경기가 없습니다.')
            embed.add_field(name='배그를 한판이라도 해주세요', value='듀오 경기 전적이 없습니다..', inline=False)
            await client.send_message(channel, embed=embed)

        else:
            duoRat1 = duoRecord1.find("span", {"class": "value"})
            duoRat = duoRat1.text.strip()  # ----레이팅----
            duoRank1 = duoRecord1.find("p", {"class": "grade-name"})
            duoRank = duoRank1.text.strip()  # ----등급----
            print(duoRank)
            embed.add_field(name='레이팅', value=duoRat, inline=False)
            embed.add_field(name='등급', value=duoRank, inline=False)

            duoStat = duoCenter1.find("div", {"class": "stats"})

            duoKD1 = duoStat.find("div", {"class": "kd stats-item stats-top-graph"})
            duoKD2 = duoKD1.find("p", {"class": "value"})
            duoKD = duoKD2.text.strip()  # ----킬뎃----
            duoKdSky1 = duoStat.find("span", {"class": "top"})
            duoKdSky = duoKdSky1.text.strip()  # ----킬뎃 상위?%----
            print(duoKD)
            print(duoKdSky)
            embed.add_field(name='킬뎃,킬뎃상위', value=duoKD + " " + duoKdSky, inline=False)

            duoWinRat1 = duoStat.find("div", {"class": "winratio stats-item stats-top-graph"})
            duoWinRat2 = duoWinRat1.find("p", {"class": "value"})
            duoWinRat = duoWinRat2.text.strip()  # ----승률----
            duoWinRatSky1 = duoWinRat1.find("span", {"class": "top"})
            duoWinRatSky = duoWinRatSky1.text.strip()  # ----승률 상위?%----
            print(duoWinRat)
            print(duoWinRatSky)
            embed.add_field(name='승률,승률상위', value=duoWinRat + " " + duoWinRatSky, inline=False)

            duoHead1 = duoStat.find("div", {"class": "headshots"})
            duoHead2 = duoHead1.find("p", {"class": "value"})
            duoHead = duoHead2.text.strip()  # ----헤드샷----
            duoHeadSky1 = duoHead1.find("span", {"class": "top"})
            duoHeadSky = duoHeadSky1.text.strip()  # ----헤드샷 상위?%----
            print(duoHead)
            print(duoHeadSky)
            embed.add_field(name='헤드샷,헤드샷상위', value=duoHead + " " + duoHeadSky, inline=False)
            await client.send_message(channel, embed=embed)

    if message.content.startswith("!배그스쿼드"):

        learn = message.content.split(" ")
        location = learn[1]
        enc_location = urllib.parse.quote(location)
        url = "https://dak.gg/profile/" + enc_location
        html = urllib.request.urlopen(url)
        bsObj = bs4.BeautifulSoup(html, "html.parser")
        duoCenter1 = bsObj.find("section", {"class": "squad modeItem"})
        duoRecord1 = duoCenter1.find("div", {"class": "overview"})
        duoRecord = duoRecord1.text.strip()  # ----기록이없습니다 문구----
        print(duoRecord)
        channel = message.channel
        embed = discord.Embed(
            title='배그스쿼드 정보',
            description='배그스쿼드 정보입니다.',
            colour=discord.Colour.green())
        if duoRecord == 'No record':
            print('스쿼드 경기가 없습니다.')
            embed.add_field(name='배그를 한판이라도 해주세요', value='스쿼드 경기 전적이 없습니다..', inline=False)
            await client.send_message(channel, embed=embed)

        else:
            duoRat1 = duoRecord1.find("span", {"class": "value"})
            duoRat = duoRat1.text.strip()  # ----레이팅----
            duoRank1 = duoRecord1.find("p", {"class": "grade-name"})
            duoRank = duoRank1.text.strip()  # ----등급----
            print(duoRank)
            embed.add_field(name='레이팅', value=duoRat, inline=False)
            embed.add_field(name='등급', value=duoRank, inline=False)

            duoStat = duoCenter1.find("div", {"class": "stats"})

            duoKD1 = duoStat.find("div", {"class": "kd stats-item stats-top-graph"})
            duoKD2 = duoKD1.find("p", {"class": "value"})
            duoKD = duoKD2.text.strip()  # ----킬뎃----
            duoKdSky1 = duoStat.find("span", {"class": "top"})
            duoKdSky = duoKdSky1.text.strip()  # ----킬뎃 상위?%----
            print(duoKD)
            print(duoKdSky)
            embed.add_field(name='킬뎃,킬뎃상위', value=duoKD + " " + duoKdSky, inline=False)

            duoWinRat1 = duoStat.find("div", {"class": "winratio stats-item stats-top-graph"})
            duoWinRat2 = duoWinRat1.find("p", {"class": "value"})
            duoWinRat = duoWinRat2.text.strip()  # ----승률----
            duoWinRatSky1 = duoWinRat1.find("span", {"class": "top"})
            duoWinRatSky = duoWinRatSky1.text.strip()  # ----승률 상위?%----
            print(duoWinRat)
            print(duoWinRatSky)
            embed.add_field(name='승률,승률상위', value=duoWinRat + " " + duoWinRatSky, inline=False)

            duoHead1 = duoStat.find("div", {"class": "headshots"})
            duoHead2 = duoHead1.find("p", {"class": "value"})
            duoHead = duoHead2.text.strip()  # ----헤드샷----
            duoHeadSky1 = duoHead1.find("span", {"class": "top"})
            duoHeadSky = duoHeadSky1.text.strip()  # ----헤드샷 상위?%----
            print(duoHead)
            print(duoHeadSky)
            embed.add_field(name='헤드샷,헤드샷상위', value=duoHead + " " + duoHeadSky, inline=False)
            await client.send_message(channel, embed=embed)

    if message.content.startswith("봇"):
        await client.send_message(message.channel, "뭐, 병X아")
    if message.content.startswith("씨발"):
        await client.send_message(message.channel, "욕하지마잉...미워! ㅠ_ㅠ")
    if message.content.startswith("병신"):
        await client.send_message(message.channel, "욕하지마잉...미워! ㅠ_ㅠ")
    if message.content.startswith("미친"):
        await client.send_message(message.channel, "욕하지마잉...미워! ㅠ_ㅠ")
        
access_token = os.environ["BOT_TOKEN"]
client.run(access_token)
